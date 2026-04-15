"""
RCM ERP — FastAPI Backend
Uruchomienie: uvicorn main:app --host 0.0.0.0 --port 8000 --reload
Dostęp z sieci lokalnej: http://192.168.1.15:8000
"""
import pathlib
import random
import uuid
from datetime import datetime, date, timezone

def _now():
    return datetime.now(timezone.utc).replace(tzinfo=None)
from typing import List, Optional

from fastapi import FastAPI, Depends, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, Response, StreamingResponse
from pdf_gen import generate_arkusz_pdf
from sqlalchemy.orm import Session
from sqlalchemy import create_engine, func, text
from sqlalchemy.orm import sessionmaker

from models import (
    Base, Order, OrderOperation, Quote, MaterialRequest,
    QualityCard, ProductTemplate, ConstraintRule, PriceHistory,
    ParameterRequest, StockMovement, ComponentContainer,
    Setting, OrderStatus, TriageBranch, OrderAttachment, init_db
)
from schemas import (
    OrderCreate, OrderOut, TriageResponse,
    QuoteCreate, QuoteZaporCreate, QuoteOut,
    QuoteStructuredCreate,
    AttachmentOut,
    MaterialRequestCreate, MaterialRequestOut,
    QualityCardCreate, QualityCardOut,
    ParameterRequestCreate, ParameterRequestAnswer, ParameterRequestOut,
    TemplateCreate, TemplateOut,
    AnalyticsSummary, RevenueMonth, TopClient, OverdueOrder,
)
from triage import run_triage, TriageInput

# ─── Konfiguracja bazy ────────────────────────────────────────────────────────
DATABASE_URL = "sqlite:///./rcm_erp.db"
engine = create_engine(DATABASE_URL, connect_args={"check_same_thread": False})
SessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
Base.metadata.create_all(bind=engine)  # tworzy tabele jeśli nie istnieją

# ─── Idempotentna migracja kolumn v2 dla istniejących baz ────────────────────
def _ensure_quote_v2_columns(eng) -> None:
    """
    Idempotentna migracja — dodaje brakujące kolumny do istniejących tabel.
    SQLite obsługuje tylko ALTER TABLE ADD COLUMN — bez ryzyka utraty danych.
    Uruchamiana przy każdym starcie serwera.
    """
    # Kolumny dodane w v2 do tabeli quotes
    quote_cols = [
        ("processes_json",     "TEXT DEFAULT '[]'"),
        ("weight_kg",          "REAL DEFAULT 0"),
        ("weight_rate_pln_kg", "REAL DEFAULT 15"),
        ("welding_hours",      "REAL DEFAULT 0"),
        ("weight_netto_kg",    "REAL DEFAULT 0"),
        ("weight_brutto_kg",   "REAL DEFAULT 0"),
        ("estimate_version",   "TEXT DEFAULT 'v1'"),
    ]
    # Kolumny dodane do tabeli orders po pierwszym deploymencie
    order_cols = [
        ("order_type",     "TEXT DEFAULT 'remont'"),
        ("sop_name",       "TEXT"),
        ("description",    "TEXT"),
        ("requires_visit", "INTEGER DEFAULT 0"),
        ("quantity",       "INTEGER DEFAULT 1"),
        ("is_defence",     "INTEGER DEFAULT 0"),
    ]
    with eng.connect() as conn:
        existing_quotes = {row[1] for row in conn.execute(text("PRAGMA table_info(quotes)"))}
        for col_name, col_def in quote_cols:
            if col_name not in existing_quotes:
                conn.execute(text(f"ALTER TABLE quotes ADD COLUMN {col_name} {col_def}"))

        existing_orders = {row[1] for row in conn.execute(text("PRAGMA table_info(orders)"))}
        for col_name, col_def in order_cols:
            if col_name not in existing_orders:
                conn.execute(text(f"ALTER TABLE orders ADD COLUMN {col_name} {col_def}"))

        conn.commit()

_ensure_quote_v2_columns(engine)

# ─── Katalog dla załączników ──────────────────────────────────────────────────
UPLOAD_ROOT = pathlib.Path(__file__).parent / "uploads"
UPLOAD_ROOT.mkdir(exist_ok=True)
MAX_UPLOAD_BYTES = 50 * 1024 * 1024   # 50 MB limit

# ─── FastAPI app ──────────────────────────────────────────────────────────────
app = FastAPI(
    title="RCM ERP",
    description="System CPQ/ERP dla RCM Sp. z o.o. — Gołdap",
    version="0.1.0",
)

# Serwowanie załączników statycznie — /uploads/{order_id}/{filename}
app.mount("/uploads", StaticFiles(directory=str(UPLOAD_ROOT)), name="uploads")

# CORS — pozwala frontendowi Vue na localhost lub 192.168.x.x łączyć się z API
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # zawęzić po wdrożeniu do adresu IP biura
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Dependency — sesja bazy ──────────────────────────────────────────────────
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


# ─── Helper: generowanie numeru zlecenia ──────────────────────────────────────
def _generate_order_number(db: Session) -> str:
    """Format: {liczba_porządkowa}/{rok}, np. "23/2026"."""
    year = date.today().year
    count = db.query(func.count(Order.id)).scalar() + 1
    return f"{count}/{year}"


# =============================================================================
# HEALTH CHECK
# =============================================================================
@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return Response(status_code=204)  # no content — ucisza spam w logach


@app.get("/api/health")
def health():
    return {"status": "ok", "system": "RCM ERP"}


# =============================================================================
# ORDERS — Zlecenia Wewnętrzne
# =============================================================================

@app.post("/api/orders", response_model=OrderOut, status_code=201)
def create_order(payload: OrderCreate, db: Session = Depends(get_db)):
    """Biuro tworzy nowe zlecenie przez Wizarda (Step 1-3)."""
    order = Order(
        order_number    = _generate_order_number(db),
        client          = payload.client,
        deadline        = payload.deadline,
        material        = payload.material,
        has_drawing     = payload.has_drawing,
        notes           = payload.notes,
        purpose         = payload.purpose,
        estimated_value = payload.estimated_value,
        order_type      = payload.order_type,
        description     = payload.description,
        requires_visit  = payload.requires_visit,
        template_id     = payload.template_id,
        quantity        = payload.quantity,
        is_defence      = payload.is_defence,
        status          = OrderStatus.draft,
    )
    db.add(order)
    db.commit()
    db.refresh(order)
    return order


@app.get("/api/orders", response_model=List[OrderOut])
def list_orders(
    status: Optional[str] = None,
    branch: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Lista wszystkich zleceń z opcjonalnym filtrem po statusie/gałęzi."""
    q = db.query(Order)
    if status:
        q = q.filter(Order.status == status)
    if branch:
        q = q.filter(Order.triage_branch == branch)
    return q.order_by(Order.created_at.desc()).all()


@app.get("/api/orders/{order_id}", response_model=OrderOut)
def get_order(order_id: int, db: Session = Depends(get_db)):
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")
    return order


@app.delete("/api/orders/{order_id}", status_code=204)
def delete_order(order_id: int, db: Session = Depends(get_db)):
    """Soft-delete: ustawia status='cancelled'. Nie usuwa z bazy."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")
    order.status = "cancelled"
    db.commit()
    return Response(status_code=204)


# =============================================================================
# TRIAGE — silnik routingu (Odrzut / Standard / Niestandard)
# =============================================================================

@app.post("/api/orders/{order_id}/triage", response_model=TriageResponse)
def triage_order(order_id: int, db: Session = Depends(get_db)):
    """
    Uruchamia Triage Engine dla zlecenia.
    Aktualizuje order.status i order.triage_branch w bazie.
    """
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")

    # Zbuduj obiekt wejściowy dla Triage Engine
    triage_input = TriageInput(
        client          = order.client,
        material        = order.material or "",
        deadline_days   = (order.deadline - date.today()).days if order.deadline else 999,
        has_drawing     = order.has_drawing,
        order_type      = order.order_type or "remont",
        sop_name        = order.sop_name,        # Fix #1: nie None!
        template_id     = order.template_id,     # Fix #5: przekazać wprost
        estimated_value = float(order.estimated_value or 0),
    )

    result = run_triage(triage_input, db)

    # Zapisz wynik triage w bazie (w tym dopasowany template_id)
    order.triage_branch = result.branch
    order.status = {
        "odrzut":      OrderStatus.rejected,
        "standard":    OrderStatus.standard,
        "niestandard": OrderStatus.niestandard,
    }[result.branch]
    if result.template_id and not order.template_id:
        order.template_id = result.template_id  # zapisz dopasowany szablon SOP
    db.commit()

    return TriageResponse(
        branch      = result.branch,
        message     = result.message,
        template_id = result.template_id,
        rule_name   = result.rule_name,
        warnings    = result.warnings or [],
    )


# =============================================================================
# QUOTES — Wycena
# =============================================================================

@app.post("/api/orders/{order_id}/quote", response_model=QuoteOut, status_code=201)
def create_quote(order_id: int, payload: QuoteCreate, db: Session = Depends(get_db)):
    """Ręczna wycena przez Technologa (gałąź Niestandard)."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")

    # Pobierz stawkę robocizny z bazy — Fix #3: nie hardkod
    rate_setting = db.get(Setting, "labor_rate_pln")
    labor_rate   = float(rate_setting.value) if rate_setting else 90.0
    # Oblicz cenę: (material + robocizna) × (1 + overhead) × (1 + marża)
    labor_cost  = payload.labor_hours * labor_rate
    subtotal    = (payload.material_cost + labor_cost) * (1 + payload.overhead_pct)
    total_net   = subtotal * (1 + payload.margin_pct)

    quote = Quote(
        order_id      = order_id,
        line_items    = payload.line_items,
        labor_hours   = payload.labor_hours,
        material_cost = payload.material_cost,
        overhead_pct  = payload.overhead_pct,
        margin_pct    = payload.margin_pct,
        total_net     = round(total_net, 2),
        is_zapor      = False,
    )
    db.add(quote)
    # Zlecenie niestandard → quoted po zapisaniu wyceny
    if order.status == OrderStatus.niestandard:
        order.status = OrderStatus.quoted
    db.commit()
    db.refresh(quote)
    return quote


@app.post("/api/orders/{order_id}/quote/zapor", response_model=QuoteOut, status_code=201)
def create_zapor_quote(order_id: int, payload: QuoteZaporCreate, db: Session = Depends(get_db)):
    """
    Zaporowa marża — Technolog klika jeden przycisk.
    Cena = material × godziny × mnożnik losowy (3.0–4.5).
    Klient albo zapłaci dużo, albo sam odpadnie — obie opcje są dobre dla zakładu.
    """
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")

    # Mnożnik zaporowy — losowy z przedziału [3.0, 4.5]
    multiplier = random.uniform(3.0, 4.5)
    total_net  = payload.material_cost * payload.hours_estimate * multiplier

    quote = Quote(
        order_id      = order_id,
        line_items    = [],
        labor_hours   = payload.hours_estimate,
        material_cost = payload.material_cost,
        overhead_pct  = 0.0,
        margin_pct    = round(multiplier - 1, 4),   # dla widoku Dyrektora
        total_net     = round(total_net, 2),
        is_zapor      = True,
    )
    db.add(quote)
    # Zlecenie niestandard → quoted po zapisaniu wyceny
    if order.status == OrderStatus.niestandard:
        order.status = OrderStatus.quoted
    db.commit()
    db.refresh(quote)
    return quote


@app.get("/api/orders/{order_id}/quote", response_model=QuoteOut)
def get_quote(order_id: int, db: Session = Depends(get_db)):
    quote = db.query(Quote).filter(Quote.order_id == order_id).first()
    if not quote:
        raise HTTPException(status_code=404, detail="Brak wyceny dla tego zlecenia")
    return quote


@app.post("/api/orders/{order_id}/quote/structured", response_model=QuoteOut, status_code=201)
def create_structured_quote(
    order_id: int, payload: QuoteStructuredCreate, db: Session = Depends(get_db)
):
    """
    Strukturalna wycena technologa wg jego formuły (v2).
    Formuła: procesy + materiał + waga×stawka + spawanie×stawka + robocizna_extra
    Następnie: ×(1+overhead) ×(1+marża)
    """
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")

    rate_setting = db.get(Setting, "labor_rate_pln")
    labor_rate   = float(rate_setting.value) if rate_setting else 90.0

    proc_total    = sum(p.cost for p in payload.processes)
    weight_total  = payload.weight_kg * payload.weight_rate_pln_kg
    welding_total = payload.welding_hours * labor_rate
    extra_labor   = payload.labor_hours * labor_rate
    base          = proc_total + payload.material_cost + weight_total + welding_total + extra_labor
    subtotal      = base * (1 + payload.overhead_pct)
    total_net     = subtotal * (1 + payload.margin_pct)

    # Upsert — jeden Quote na zlecenie (zastępuje poprzednią wycenę)
    quote = db.query(Quote).filter(Quote.order_id == order_id).first()
    if not quote:
        quote = Quote(order_id=order_id)
        db.add(quote)

    quote.line_items         = []
    quote.processes_json     = [p.model_dump() for p in payload.processes]
    quote.material_cost      = payload.material_cost
    quote.weight_kg          = payload.weight_kg
    quote.weight_rate_pln_kg = payload.weight_rate_pln_kg
    quote.welding_hours      = payload.welding_hours
    quote.weight_netto_kg    = payload.weight_netto_kg
    quote.weight_brutto_kg   = payload.weight_brutto_kg
    quote.labor_hours        = payload.labor_hours + payload.welding_hours
    quote.overhead_pct       = payload.overhead_pct
    quote.margin_pct         = payload.margin_pct
    quote.total_net          = round(total_net, 2)
    quote.is_zapor           = False
    quote.estimate_version   = "v2"

    # Zlecenie niestandard → quoted
    if order.status == OrderStatus.niestandard:
        order.status = OrderStatus.quoted

    db.commit()
    db.refresh(quote)
    return quote


@app.post("/api/orders/{order_id}/confirm", response_model=OrderOut)
def confirm_order(order_id: int, db: Session = Depends(get_db)):
    """Biuro zatwierdza wycenę Technologa → zlecenie trafia do produkcji."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")
    if order.status != OrderStatus.quoted:
        raise HTTPException(status_code=409, detail="Tylko zlecenia w statusie 'quoted' można zatwierdzić")
    order.status = OrderStatus.in_production
    db.commit()
    db.refresh(order)
    return order


@app.post("/api/orders/{order_id}/start", response_model=OrderOut)
def start_order(order_id: int, db: Session = Depends(get_db)):
    """Technolog rozpoczyna pracę: in_production → w_trakcie."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")
    if order.status != OrderStatus.in_production:
        raise HTTPException(status_code=409, detail="Tylko zlecenia 'in_production' można rozpocząć")
    order.status = OrderStatus.w_trakcie
    db.commit()
    db.refresh(order)
    return order


@app.post("/api/orders/{order_id}/complete", response_model=OrderOut)
def complete_order(order_id: int, db: Session = Depends(get_db)):
    """Technolog kończy pracę: w_trakcie → gotowe (czeka na odbiór)."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")
    if order.status != OrderStatus.w_trakcie:
        raise HTTPException(status_code=409, detail="Tylko zlecenia 'w_trakcie' można oznaczyć jako gotowe")
    order.status = OrderStatus.gotowe
    db.commit()
    db.refresh(order)
    return order


@app.post("/api/orders/{order_id}/deliver", response_model=OrderOut)
def deliver_order(order_id: int, db: Session = Depends(get_db)):
    """Biuro wydaje zlecenie klientowi: gotowe → wydane."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")
    if order.status != OrderStatus.gotowe:
        raise HTTPException(status_code=409, detail="Tylko zlecenia 'gotowe' można oznaczyć jako wydane")
    order.status = OrderStatus.wydane
    db.commit()
    db.refresh(order)
    return order


@app.post("/api/orders/{order_id}/save-as-template", response_model=TemplateOut, status_code=201)
def save_order_as_template(order_id: int, payload: dict, db: Session = Depends(get_db)):
    """
    Technolog zapisuje niestandard jako szablon SOP — jeśli zlecenie się powtórzy,
    następnym razem trafi do gałęzi Standard automatycznie.
    """
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")
    quote = db.query(Quote).filter(Quote.order_id == order_id).first()

    # Wyprowadź operacje z processes_json (v2) lub pusty
    ops = []
    if quote and quote.processes_json:
        ops = [{"op": p["name"], "hours": 0, "cost": p["cost"]} for p in quote.processes_json]

    # Materiał z wagi jeśli dostępny
    mats = []
    if quote and quote.weight_kg:
        mats = [{"mat": order.material or "Materiał", "qty_kg": float(quote.weight_kg), "unit": "kg"}]

    tmpl = ProductTemplate(
        name               = payload.get("name") or f"Z zlecenia {order.order_number}",
        category           = payload.get("category") or order.order_type or "remont",
        operations_json    = ops,
        materials_json     = mats,
        instruction_blocks = [],
        machines_json      = [],
        base_price_pln     = float(quote.total_net) if quote and quote.total_net else None,
        margin_pct         = float(quote.margin_pct or 0.25) if quote else 0.25,
    )
    db.add(tmpl)
    db.commit()
    db.refresh(tmpl)
    return tmpl


# =============================================================================
# ATTACHMENTS — Załączniki (rysunki, dokumentacja techniczna)
# =============================================================================

@app.post("/api/orders/{order_id}/attachments", response_model=AttachmentOut, status_code=201)
async def upload_attachment(
    order_id: int,
    file: UploadFile = File(...),
    uploaded_by: str = Form(default="technolog"),
    db: Session = Depends(get_db),
):
    """Wgraj plik (rysunek, PDF) do zlecenia. Max 50 MB."""
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")

    data = await file.read()
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(status_code=413, detail="Plik za duży (max 50 MB)")

    # Folder dla zlecenia, plik z losowym prefiksem (zapobiega kolizjom nazw)
    order_dir = UPLOAD_ROOT / str(order_id)
    order_dir.mkdir(exist_ok=True)
    safe_name = f"{uuid.uuid4().hex[:8]}_{file.filename}"
    dest      = order_dir / safe_name
    dest.write_bytes(data)

    att = OrderAttachment(
        order_id    = order_id,
        filename    = file.filename,
        stored_path = f"uploads/{order_id}/{safe_name}",
        size_bytes  = len(data),
        mime_type   = file.content_type,
        uploaded_by = uploaded_by,
    )
    db.add(att)
    db.commit()
    db.refresh(att)
    return att


@app.get("/api/orders/{order_id}/attachments", response_model=List[AttachmentOut])
def list_attachments(order_id: int, db: Session = Depends(get_db)):
    """Lista załączników dla zlecenia."""
    return db.query(OrderAttachment).filter(OrderAttachment.order_id == order_id).all()


@app.delete("/api/attachments/{att_id}", status_code=204)
def delete_attachment(att_id: int, db: Session = Depends(get_db)):
    """Usuń załącznik — kasuje plik z dysku i rekord z bazy."""
    att = db.get(OrderAttachment, att_id)
    if not att:
        raise HTTPException(status_code=404, detail="Załącznik nie znaleziony")
    pathlib.Path(att.stored_path).unlink(missing_ok=True)   # usuń plik (cicho jeśli już nie ma)
    db.delete(att)
    db.commit()
    return Response(status_code=204)


# =============================================================================
# MATERIAL REQUESTS — Zapotrzebowanie Materiałowe
# =============================================================================

@app.post("/api/orders/{order_id}/materials", response_model=MaterialRequestOut, status_code=201)
def create_material_request(
    order_id: int, payload: MaterialRequestCreate, db: Session = Depends(get_db)
):
    req = MaterialRequest(
        order_id    = order_id,
        client      = payload.client,
        materials   = payload.materials,
        extra_notes = payload.extra_notes,
        priority    = payload.priority,
        status      = "pending",
    )
    db.add(req)
    db.commit()
    db.refresh(req)
    return req


@app.get("/api/orders/{order_id}/materials", response_model=MaterialRequestOut)
def get_material_request(order_id: int, db: Session = Depends(get_db)):
    req = db.query(MaterialRequest).filter(MaterialRequest.order_id == order_id).first()
    if not req:
        raise HTTPException(status_code=404, detail="Brak zapotrzebowania materiałowego")
    return req


# =============================================================================
# QUALITY CARDS — Karta Kontrolna (po każdym etapie!)
# =============================================================================

@app.post("/api/orders/{order_id}/quality", response_model=QualityCardOut, status_code=201)
def add_quality_card(
    order_id: int, payload: QualityCardCreate, db: Session = Depends(get_db)
):
    card = QualityCard(
        order_id        = order_id,
        operation_id    = payload.operation_id,
        stage_name      = payload.stage_name,
        check_linear    = payload.check_linear,
        check_geometric = payload.check_geometric,
        check_surface   = payload.check_surface,
        passed          = payload.passed,
        checked_by      = payload.checked_by,
        checked_at      = _now(),
    )
    db.add(card)
    db.commit()
    db.refresh(card)
    return card


@app.get("/api/orders/{order_id}/quality", response_model=List[QualityCardOut])
def get_quality_cards(order_id: int, db: Session = Depends(get_db)):
    return db.query(QualityCard).filter(QualityCard.order_id == order_id).all()


# =============================================================================
# PARAMETER REQUESTS — "Zapytaj o parametry" (Technolog → Biuro)
# =============================================================================

@app.post("/api/orders/{order_id}/params", response_model=ParameterRequestOut, status_code=201)
def ask_for_params(
    order_id: int, payload: ParameterRequestCreate, db: Session = Depends(get_db)
):
    """Technolog wysyła pytanie do Biuro — bez telefonu."""
    req = ParameterRequest(
        order_id      = order_id,
        question_text = payload.question_text,
        status        = "pending",
    )
    db.add(req)
    db.commit()
    db.refresh(req)
    return req


@app.patch("/api/params/{param_id}/answer", response_model=ParameterRequestOut)
def answer_param_request(
    param_id: int, payload: ParameterRequestAnswer, db: Session = Depends(get_db)
):
    """Biuro odpowiada na pytanie Technologa."""
    req = db.get(ParameterRequest, param_id)
    if not req:
        raise HTTPException(status_code=404, detail="Pytanie nie znalezione")

    req.answer_text  = payload.answer_text
    req.status       = "answered"
    req.answered_at  = _now()
    db.commit()
    db.refresh(req)
    return req


@app.get("/api/orders/{order_id}/params", response_model=List[ParameterRequestOut])
def get_param_requests(order_id: int, db: Session = Depends(get_db)):
    return db.query(ParameterRequest).filter(ParameterRequest.order_id == order_id).all()


@app.get("/api/params", response_model=List[ParameterRequestOut])
def get_all_param_requests(status: Optional[str] = None, db: Session = Depends(get_db)):
    """Wszystkie pytania Technologa — dla Biuro (filtr po statusie: pending/answered)."""
    q = db.query(ParameterRequest)
    if status:
        q = q.filter(ParameterRequest.status == status)
    return q.order_by(ParameterRequest.asked_at.desc()).all()


# =============================================================================
# PRODUCT TEMPLATES — Katalog SOP (Technolog zarządza)
# =============================================================================

@app.get("/api/templates", response_model=List[TemplateOut])
def list_templates(category: Optional[str] = None, db: Session = Depends(get_db)):
    q = db.query(ProductTemplate).filter(ProductTemplate.is_active == True)
    if category:
        q = q.filter(ProductTemplate.category == category)
    return q.all()


@app.post("/api/templates", response_model=TemplateOut, status_code=201)
def create_template(payload: TemplateCreate, db: Session = Depends(get_db)):
    """Technolog tworzy szablon SOP raz → Biuro używa wiele razy."""
    tmpl = ProductTemplate(
        name               = payload.name,
        category           = payload.category,
        operations_json    = payload.operations_json,
        materials_json     = payload.materials_json,
        instruction_blocks = payload.instruction_blocks,
        machines_json      = payload.machines_json,
        base_price_pln     = payload.base_price_pln,
        margin_pct         = payload.margin_pct,
    )
    db.add(tmpl)
    db.commit()
    db.refresh(tmpl)
    return tmpl


@app.delete("/api/templates/{template_id}", status_code=204)
def archive_template(template_id: int, db: Session = Depends(get_db)):
    """
    Soft delete szablonu — ustawia is_active=False, NIE usuwa z bazy.
    Stare zlecenia z tym template_id nadal działają i generują PDF.
    Szablon znika z listy u Biuro, ale historia jest nienaruszona.
    """
    tmpl = db.get(ProductTemplate, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="Szablon nie znaleziony")
    if not tmpl.is_active:
        raise HTTPException(status_code=409, detail="Szablon już zarchiwizowany")
    tmpl.is_active = False
    db.commit()


@app.patch("/api/templates/{template_id}/restore", response_model=TemplateOut)
def restore_template(template_id: int, db: Session = Depends(get_db)):
    """Przywraca zarchiwizowany szablon (odwrócenie soft delete)."""
    tmpl = db.get(ProductTemplate, template_id)
    if not tmpl:
        raise HTTPException(status_code=404, detail="Szablon nie znaleziony")
    tmpl.is_active = True
    db.commit()
    db.refresh(tmpl)
    return tmpl


# =============================================================================
# ANALYTICS — Dashboard Dyrektora (tylko odczyt)
# =============================================================================

@app.get("/api/analytics", response_model=AnalyticsSummary)
def get_analytics(db: Session = Depends(get_db)):
    total         = db.query(func.count(Order.id)).scalar()
    odrzut_count  = db.query(func.count(Order.id)).filter(Order.triage_branch == "odrzut").scalar()
    standard      = db.query(func.count(Order.id)).filter(Order.triage_branch == "standard").scalar()
    niestandard   = db.query(func.count(Order.id)).filter(Order.triage_branch == "niestandard").scalar()
    # Liczymy wszystkie "aktywne" zlecenia produkcyjne (in_production + w_trakcie + gotowe)
    in_prod       = db.query(func.count(Order.id)).filter(
        Order.status.in_(["in_production", "w_trakcie", "gotowe"])
    ).scalar()
    done          = db.query(func.count(Order.id)).filter(
        Order.status.in_(["done", "wydane"])
    ).scalar()
    avg_margin    = db.query(func.avg(Quote.margin_pct)).scalar()

    # Przychód za ostatnie 6 miesięcy — suma wycen zatwierdzonych zleceń
    revenue_rows = db.execute(text("""
        SELECT strftime('%Y-%m', o.created_at) as month,
               COUNT(o.id) as cnt,
               COALESCE(SUM(q.total_net), 0) as revenue
        FROM orders o
        LEFT JOIN quotes q ON q.order_id = o.id
        WHERE o.status IN ('w_trakcie','gotowe','wydane','done','in_production')
          AND o.created_at >= date('now', '-6 months')
        GROUP BY month
        ORDER BY month ASC
    """)).fetchall()
    revenue_by_month = [
        RevenueMonth(month=r[0], orders=r[1], revenue_pln=float(r[2]))
        for r in revenue_rows
    ]

    # Top 5 klientów po liczbie zleceń i przychodzie
    top_rows = db.execute(text("""
        SELECT o.client,
               COUNT(o.id) as cnt,
               COALESCE(SUM(q.total_net), 0) as revenue
        FROM orders o
        LEFT JOIN quotes q ON q.order_id = o.id
        WHERE o.status NOT IN ('rejected')
        GROUP BY o.client
        ORDER BY cnt DESC, revenue DESC
        LIMIT 5
    """)).fetchall()
    top_clients = [
        TopClient(client=r[0], orders=r[1], revenue_pln=float(r[2]))
        for r in top_rows
    ]

    # Przeterminowane zlecenia — termin minął, a status nie jest końcowy
    today = date.today().isoformat()
    overdue_rows = db.execute(text(f"""
        SELECT id, order_number, client, status, deadline
        FROM orders
        WHERE deadline < '{today}'
          AND status NOT IN ('done','wydane','rejected')
        ORDER BY deadline ASC
    """)).fetchall()
    overdue_orders = [
        OverdueOrder(
            id=r[0], order_number=r[1], client=r[2],
            status=r[3], deadline=r[4]
        )
        for r in overdue_rows
    ]

    return AnalyticsSummary(
        total_orders          = total or 0,
        odrzut_count          = odrzut_count or 0,
        odrzut_pct            = round((odrzut_count / total * 100) if total else 0, 1),
        standard_count        = standard or 0,
        niestandard_count     = niestandard or 0,
        avg_margin_pct        = round(float(avg_margin) * 100, 1) if avg_margin else None,
        orders_in_production  = in_prod or 0,
        orders_done           = done or 0,
        revenue_by_month      = revenue_by_month,
        top_clients           = top_clients,
        overdue_orders        = overdue_orders,
    )


@app.get("/api/export/xlsx")
def export_xlsx(db: Session = Depends(get_db)):
    """Eksport wszystkich zleceń do pliku Excel (dla Dyrektora)."""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    orders = db.query(Order).order_by(Order.created_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Zlecenia RCM"

    # Nagłówek
    headers = ["Nr", "Klient", "Status", "Gałąź", "Termin", "Wartość netto (PLN)", "Utworzono"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1A3A5C")  # --rcm-blue
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Dane
    for row_idx, o in enumerate(orders, 2):
        quote = db.query(Quote).filter(Quote.order_id == o.id).first()
        ws.cell(row=row_idx, column=1, value=o.order_number)
        ws.cell(row=row_idx, column=2, value=o.client)
        ws.cell(row=row_idx, column=3, value=o.status.value if o.status else "")
        ws.cell(row=row_idx, column=4, value=o.triage_branch or "")
        ws.cell(row=row_idx, column=5, value=o.deadline.isoformat() if o.deadline else "")
        ws.cell(row=row_idx, column=6, value=float(quote.total_price_net) if quote and quote.total_price_net else 0)
        ws.cell(row=row_idx, column=7, value=o.created_at.strftime("%Y-%m-%d") if o.created_at else "")

    # Szerokości kolumn
    for col, width in zip(range(1, 8), [12, 25, 15, 12, 12, 20, 12]):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=zlecenia_rcm.xlsx"},
    )


# =============================================================================
# HARMONOGRAM — Lista Zleceń (Dokument 4 z kartki)
# Dyrektor i Technolog widzą wszystkie zlecenia posortowane po terminie
# Brak osobnej tabeli — to jest widok SELECT orders JOIN operations
# =============================================================================

@app.get("/api/harmonogram")
def get_harmonogram(db: Session = Depends(get_db)):
    """Zwraca wszystkie aktywne zlecenia posortowane po deadlinie."""
    orders = (
        db.query(Order)
        .filter(Order.status.notin_(["rejected", "done"]))
        .order_by(Order.deadline.asc())
        .all()
    )
    return [
        {
            "id":           o.id,
            "order_number": o.order_number,
            "client":       o.client,
            "status":       o.status,
            "deadline":     o.deadline.isoformat() if o.deadline else None,
            "branch":       o.triage_branch,
        }
        for o in orders
    ]


# =============================================================================
# SETTINGS — Ustawienia systemowe (stawka robocizny, itp.)
# =============================================================================

@app.get("/api/settings")
def get_settings(db: Session = Depends(get_db)):
    """Lista wszystkich ustawień systemowych."""
    return db.query(Setting).all()


@app.patch("/api/settings/{key}")
def update_setting(key: str, payload: dict, db: Session = Depends(get_db)):
    """Zmień wartość ustawienia (np. stawkę robocizny) bez restartu serwera."""
    setting = db.get(Setting, key)
    if not setting:
        raise HTTPException(status_code=404, detail=f"Ustawienie '{key}' nie istnieje")
    setting.value = str(payload.get("value", setting.value))
    db.commit()
    return {"key": key, "value": setting.value}


# =============================================================================
# PDF — Arkusz Zlecenia (zastępuje ręczne Excele CNC/Montaż)
# =============================================================================

@app.get("/api/orders/{order_id}/pdf")
def get_order_pdf(order_id: int, db: Session = Depends(get_db)):
    """
    Generuje PDF Arkusza Zlecenia dla operatorów CNC/Montaż.
    Zawiera: dane zlecenia, instrukcję SOP (instruction_blocks), Kartę Kontrolną.
    """
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")

    template = db.get(ProductTemplate, order.template_id) if order.template_id else None
    quote    = db.query(Quote).filter(Quote.order_id == order_id).first()

    try:
        from pdf_gen import get_content_type
        pdf_bytes    = generate_arkusz_pdf(order, template, quote)
        content_type = get_content_type(order)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Błąd generowania PDF: {e}")

    ext      = "pdf" if "pdf" in content_type else "html"
    filename = f"Arkusz_{(order.order_number or str(order_id)).replace('/', '-')}.{ext}"
    return Response(
        content=pdf_bytes,
        media_type=content_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.get("/api/orders/{order_id}/oferta")
def get_order_oferta(order_id: int, db: Session = Depends(get_db)):
    """
    Generuje PDF Oferty Handlowej dla klienta (Biuro → klient email/WhatsApp).
    Zawiera: cena netto + brutto, termin, warunki płatności.
    NIE zawiera: stawek, robocizny, marży — to tajemnica handlowa.
    """
    order = db.get(Order, order_id)
    if not order:
        raise HTTPException(status_code=404, detail="Zlecenie nie znalezione")

    template = db.get(ProductTemplate, order.template_id) if order.template_id else None
    quote    = db.query(Quote).filter(Quote.order_id == order_id).first()

    try:
        from pdf_gen import generate_oferta_pdf, get_content_type
        pdf_bytes    = generate_oferta_pdf(order, template, quote)
        content_type = get_content_type(order)
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Błąd generowania Oferty: {e}")

    ext      = "pdf" if "pdf" in content_type else "html"
    filename = f"Oferta_{(order.order_number or str(order_id)).replace('/', '-')}.{ext}"
    return Response(
        content=pdf_bytes,
        media_type=content_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =============================================================================
# FRONTEND — serwowanie pliku Vue
# =============================================================================

@app.get("/", response_class=FileResponse)
def serve_frontend():
    """Serwuje główny plik Vue 3 CDN — jeden plik dla wszystkich ról."""
    return FileResponse("../frontend/app.html")

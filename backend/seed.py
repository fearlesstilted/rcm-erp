"""
Seed danych — parsuje ceny/*.xlsx i wypełnia bazę danymi startowymi.
Uruchomienie: python seed.py

Parsuje:
  - Lista zleceń usługi.xlsx  → price_history
  - Lista zleceń usługi_1.xlsx → price_history
  - Wstawia domyślne constraint_rules (Odrzut)
  - Wstawia przykładowe product_templates
"""
import os
import sys
from datetime import date

# Dodaj folder nadrzędny do ścieżki importu
sys.path.insert(0, os.path.dirname(__file__))

from sqlalchemy.orm import Session
import re
from models import (
    Base, ConstraintRule, ProductTemplate, PriceHistory,
    User, UserRole, Setting, ApprovedMaterial, init_db
)

CENY_DIR = os.path.join(os.path.dirname(__file__), "..", "..", "ceny")


def seed_users(db: Session):
    """Domyślni użytkownicy dla MVP."""
    if db.query(User).count() > 0:
        return  # już seededowane

    users = [
        User(name="Biuro",     role=UserRole.biuro,     pin="1111"),
        User(name="Technolog", role=UserRole.technolog,  pin="2222"),
        User(name="Dyrektor",  role=UserRole.dyrektor,   pin="3333"),
    ]
    db.add_all(users)
    db.commit()
    print("✓ Użytkownicy dodani (PIN: Biuro=1111, Technolog=2222, Dyrektor=3333)")


def seed_constraint_rules(db: Session):
    """
    Reguły Odrzutu — wstawia raz, Technolog może je edytować przez API.
    Dane z rozmów z technologiem (materiały, terminy, marża minimalna).
    """
    if db.query(ConstraintRule).count() > 0:
        return

    rules = [
        ConstraintRule(
            rule_name = "Materiał nierdzewny",
            field     = "material",
            operator  = "in",
            value     = "nierdzewka,nierdzewne,stal nierdzewna,inox,304,316",
            action    = "reject",
            message   = "Odrzut: nie pracujemy w stali nierdzewnej.",
        ),
        ConstraintRule(
            rule_name = "Materiał aluminium",
            field     = "material",
            operator  = "in",
            value     = "aluminium,al,alum",
            action    = "reject",
            message   = "Odrzut: nie pracujemy w aluminium.",
        ),
        ConstraintRule(
            rule_name = "Termin za krótki",
            field     = "deadline_days",
            operator  = "lt",
            value     = "3",
            action    = "reject",
            message   = "Odrzut: za krótki termin (minimum 3 dni robocze).",
        ),
        ConstraintRule(
            rule_name = "Zbyt mała marża",
            field     = "estimated_value",
            operator  = "lt",
            value     = "500",
            action    = "warn",   # ostrzeżenie, nie twardy odrzut
            message   = "Uwaga: szacowana wartość poniżej 500 zł — sprawdź opłacalność.",
        ),
    ]
    db.add_all(rules)
    db.commit()
    print(f"✓ {len(rules)} reguł Odrzutu wstawionych")


def seed_product_templates(db: Session):
    """
    Szablony SOP — wstawia typowe prace RCM.
    Dane wyciągnięte z nazw plików w ceny/ i KnowledgeTopKop.json.
    """
    if db.query(ProductTemplate).count() > 0:
        return

    templates = [
        ProductTemplate(
            name     = "Wymiana zęba w łyżce kopiącej",
            category = "remont",
            operations_json = [
                {"op": "SPAWANIE", "hours": 2.0, "responsible": "spawacz"},
                {"op": "SZLIFOWANIE", "hours": 0.5, "responsible": "spawacz"},
            ],
            materials_json = [
                {"mat": "Ząb do łyżki", "qty": 1, "unit": "szt"},
                {"mat": "Elektroda spawalnicza", "qty": 0.5, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Oczyścić powierzchnię z rdzy i zanieczyszczeń"},
                {"order": 2, "text": "Usunąć stary ząb metodą cięcia plazmą"},
                {"order": 3, "text": "Wspawać nowy ząb wg. pozycji fabrycznej"},
                {"order": 4, "text": "Zeszlifować spoiny do gładkości"},
                {"order": 5, "text": "Sprawdzić wymiary wg. Karty Kontrolnej"},
            ],
            machines_json = [
                {"machine": "SPAWARKA MIG", "required": True},
                {"machine": "PLAZMA", "required": True},
                {"machine": "SZLIFIERKA", "required": True},
            ],
            base_price_pln = 450.0,
            margin_pct     = 0.25,
        ),
        ProductTemplate(
            name     = "Naprawa ramy / prostowanie",
            category = "remont",
            operations_json = [
                {"op": "PROSTOWANIE", "hours": 3.0, "responsible": "ślusarz"},
                {"op": "SPAWANIE", "hours": 1.5, "responsible": "spawacz"},
            ],
            materials_json = [
                {"mat": "Blacha S355", "qty": 5, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Oczyścić powierzchnię"},
                {"order": 2, "text": "Wykonać pomiar odkształcenia"},
                {"order": 3, "text": "Prostować na prasie hydraulicznej"},
                {"order": 4, "text": "Wspawać wzmocnienia jeśli wymagane"},
                {"order": 5, "text": "Sprawdzenie wymiarów geometrycznych"},
            ],
            machines_json = [
                {"machine": "PRASA", "required": True},
                {"machine": "SPAWARKA MIG", "required": False},
            ],
            base_price_pln = 800.0,
            margin_pct     = 0.25,
        ),
        ProductTemplate(
            name     = "Wspawanie elementu / naплавanie",
            category = "remont",
            operations_json = [
                {"op": "SPAWANIE", "hours": 1.5, "responsible": "spawacz"},
            ],
            materials_json = [
                {"mat": "Elektroda / drut spawalniczy", "qty": 1, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Oczyścić powierzchnię i wspawać nowy element"},
                {"order": 2, "text": "Zeszlifować spoiny"},
                {"order": 3, "text": "Sprawdzenie wymiarów liniowych"},
            ],
            machines_json = [
                {"machine": "SPAWARKA MIG", "required": True},
            ],
            base_price_pln = 300.0,
            margin_pct     = 0.25,
        ),
        ProductTemplate(
            name     = "Cięcie plazmą CNC",
            category = "usługa",
            operations_json = [
                {"op": "CNC PLAZMA", "hours": 1.0, "responsible": "operator CNC"},
            ],
            materials_json = [],
            instruction_blocks = [
                {"order": 1, "text": "Wczytać plik DXF do systemu CNC"},
                {"order": 2, "text": "Wykonać cięcie wg. parametrów materiału"},
                {"order": 3, "text": "Sprawdzenie wymiarów liniowych po cięciu"},
            ],
            machines_json = [
                {"machine": "LASER/PLAZMA CNC", "required": True},
            ],
            base_price_pln = 200.0,
            margin_pct     = 0.30,
        ),

        # ── Szablony z prawdziwych zleceń (ceny/ folder, kwiecień 2026) ──────

        ProductTemplate(
            name     = "Wspawanie zębów do wiertła",
            category = "remont",
            # Realne dane: zlec. 18/26, ok. 544 PLN netto
            operations_json = [
                {"op": "SPAWANIE", "hours": 2.0, "responsible": "spawacz"},
                {"op": "SZLIFOWANIE", "hours": 0.5, "responsible": "spawacz"},
            ],
            materials_json = [
                {"mat": "Ząb do wiertła", "qty": 1, "unit": "szt"},
                {"mat": "Elektroda spawalnicza", "qty": 0.3, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Oczyścić powierzchnię wiertła z rdzy i błota"},
                {"order": 2, "text": "Usunąć zużyte zęby metodą cięcia plazmą"},
                {"order": 3, "text": "Dopasować i wspawać nowe zęby wg. pozycji fabrycznej"},
                {"order": 4, "text": "Zeszlifować spoiny do profilu oryginalnego"},
                {"order": 5, "text": "Sprawdzić wymiary liniowe i symetrię osadzenia"},
            ],
            machines_json = [
                {"machine": "SPAWARKA MIG", "required": True},
                {"machine": "PLAZMA", "required": True},
                {"machine": "SM", "required": True},
            ],
            base_price_pln = 435.0,   # 544 netto / 1.25 marży
            margin_pct     = 0.25,
        ),

        ProductTemplate(
            name     = "Wymiana zęba w łyżce kopiącej (zlec. 26/26)",
            category = "remont",
            # Realne dane: zlec. 26/26, data 10.04.2026, ok. 754 PLN netto
            operations_json = [
                {"op": "CIĘCIE PLAZMĄ", "hours": 0.5, "responsible": "operator CNC"},
                {"op": "SPAWANIE", "hours": 2.0, "responsible": "spawacz"},
                {"op": "SZLIFOWANIE", "hours": 0.5, "responsible": "spawacz"},
            ],
            materials_json = [
                {"mat": "Ząb do łyżki kopiącej", "qty": 1, "unit": "szt"},
                {"mat": "Elektroda spawalnicza", "qty": 0.5, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Oczyścić łyżkę z ziemi i zanieczyszczeń"},
                {"order": 2, "text": "Wyciąć stary ząb plazmą"},
                {"order": 3, "text": "Dopasować nowy ząb do gniazda"},
                {"order": 4, "text": "Wspawać ząb wg. pozycji fabrycznej (MIG)"},
                {"order": 5, "text": "Zeszlifować spoiny"},
                {"order": 6, "text": "Sprawdzenie wymiarów liniowych i geometrycznych"},
            ],
            machines_json = [
                {"machine": "PLAZMA", "required": True},
                {"machine": "SPAWARKA MIG", "required": True},
                {"machine": "SM", "required": True},
            ],
            base_price_pln = 603.0,   # 754 netto / 1.25
            margin_pct     = 0.25,
        ),

        ProductTemplate(
            name     = "Remont łyżki koparkowej (montaż + spawanie)",
            category = "remont",
            # Realne dane: zlec. 04/26, Palwod, W-2 MONTAŻ, ok. 1273 PLN netto
            operations_json = [
                {"op": "PIŁA", "hours": 0.5, "responsible": "operator"},
                {"op": "SPAWANIE", "hours": 4.0, "responsible": "spawacz"},
                {"op": "MONTAŻ", "hours": 2.0, "responsible": "ślusarz"},
                {"op": "SZLIFOWANIE", "hours": 1.0, "responsible": "spawacz"},
            ],
            materials_json = [
                {"mat": "Blacha Hardox 400", "qty": 15, "unit": "kg"},
                {"mat": "Elektroda spawalnicza", "qty": 1.5, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Oczyścić łyżkę, ocenić zakres uszkodzeń"},
                {"order": 2, "text": "Wyciąć uszkodzone elementy (PIŁA/PLAZMA)"},
                {"order": 3, "text": "Przyciąć wstawki z blachy Hardox wg. wymiarów"},
                {"order": 4, "text": "Wspawać wstawki (pełny przetop)"},
                {"order": 5, "text": "Zmontować elementy śrubowe i osie"},
                {"order": 6, "text": "Zeszlifować spoiny do gładkości"},
                {"order": 7, "text": "Sprawdzenie wymiarów geometrycznych"},
            ],
            machines_json = [
                {"machine": "PIŁA", "required": True},
                {"machine": "SPAWARKA MIG", "required": True},
                {"machine": "SM", "required": True},
                {"machine": "MONTAŻ", "required": True},
            ],
            base_price_pln = 1018.0,  # 1273 netto / 1.25
            margin_pct     = 0.25,
        ),

        ProductTemplate(
            name     = "Zbrojenie dekiel szamba (montaż CNC)",
            category = "zbrojenie",
            # Realne dane: zbrojenie dekiel szamba, ok. 3824 PLN netto
            operations_json = [
                {"op": "CNC PLAZMA", "hours": 3.0, "responsible": "operator CNC"},
                {"op": "MONTAŻ", "hours": 6.0, "responsible": "zbrojarz"},
                {"op": "SPAWANIE", "hours": 2.0, "responsible": "spawacz"},
            ],
            materials_json = [
                {"mat": "Pręt żebrowany B500SP Ø8", "qty": 50, "unit": "kg"},
                {"mat": "Pręt żebrowany B500SP Ø10", "qty": 30, "unit": "kg"},
                {"mat": "Drut wiązałkowy", "qty": 1, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Pobrać pręty wg. listy materiałowej"},
                {"order": 2, "text": "Ciąć pręty na wymiar (CNC lub PIŁA)"},
                {"order": 3, "text": "Giąć strzemiona wg. rysunku"},
                {"order": 4, "text": "Montować kosz zbrojeniowy wg. rysunku"},
                {"order": 5, "text": "Wiązać węzły drutem wiązałkowym"},
                {"order": 6, "text": "Sprawdzenie wymiarów liniowych kosza"},
                {"order": 7, "text": "Oznaczyć pojemnik numerem zlecenia"},
            ],
            machines_json = [
                {"machine": "LASER/PLAZMA CNC", "required": False},
                {"machine": "PIŁA", "required": True},
                {"machine": "GIĘTARKA", "required": True},
            ],
            base_price_pln = 3059.0,  # 3824 netto / 1.25
            margin_pct     = 0.25,
        ),

        ProductTemplate(
            name     = "Strzemiona CNC (RCM, seria)",
            category = "zbrojenie",
            # Realne dane: zlec. 19/26, Stzremiona RCM, ok. 450 PLN / seria
            operations_json = [
                {"op": "CNC GIĘTARKA", "hours": 2.0, "responsible": "operator CNC"},
            ],
            materials_json = [
                {"mat": "Pręt B500SP Ø6", "qty": 20, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Wczytać program gięcia do maszyny CNC"},
                {"order": 2, "text": "Ustawić wymiary wg. rysunku (a, b, c)"},
                {"order": 3, "text": "Wykonać serię strzemion"},
                {"order": 4, "text": "Sprawdzić wymiary pierwszej sztuki kontrolnej"},
                {"order": 5, "text": "Skompletować i oznaczyć paczkę numerem zlecenia"},
            ],
            machines_json = [
                {"machine": "CNC GIĘTARKA", "required": True},
            ],
            base_price_pln = 360.0,
            margin_pct     = 0.25,
        ),

        ProductTemplate(
            name     = "Zbrojenie piwniczki — schody / wejście",
            category = "zbrojenie",
            # Realne dane: zlec. 24/26, piwniczka schody + wejście CNC + montaż
            operations_json = [
                {"op": "CNC PLAZMA", "hours": 4.0, "responsible": "operator CNC"},
                {"op": "MONTAŻ", "hours": 8.0, "responsible": "zbrojarz"},
            ],
            materials_json = [
                {"mat": "Pręt żebrowany B500SP Ø8", "qty": 80, "unit": "kg"},
                {"mat": "Pręt żebrowany B500SP Ø12", "qty": 40, "unit": "kg"},
                {"mat": "Drut wiązałkowy", "qty": 2, "unit": "kg"},
            ],
            instruction_blocks = [
                {"order": 1, "text": "Sprawdzić rysunek — schody czy wejście (2 warianty)"},
                {"order": 2, "text": "Pobrać pręty wg. listy materiałowej BOM"},
                {"order": 3, "text": "Ciąć i giąć pręty CNC wg. programu"},
                {"order": 4, "text": "Montować kosz zbrojeniowy na stole montażowym"},
                {"order": 5, "text": "Wiązać węzły co 200mm drutem"},
                {"order": 6, "text": "Sprawdzenie wymiarów: dług. × szer. × wys."},
                {"order": 7, "text": "Umieścić w pojemniku, oznaczyć zleceniem"},
            ],
            machines_json = [
                {"machine": "PIŁA", "required": True},
                {"machine": "CNC GIĘTARKA", "required": True},
            ],
            base_price_pln = 18584.0,  # 23230 netto / 1.25 (duże zlecenie UNIBEP/DBK)
            margin_pct     = 0.25,
        ),
    ]
    db.add_all(templates)
    db.commit()
    print(f"✓ {len(templates)} szablonów produktów wstawionych")


def seed_settings(db: Session):
    """Wartości domyślne ustawień systemowych."""
    defaults = [
        Setting(key="labor_rate_pln", value="90", label="Stawka robocizny (PLN/h)"),
        Setting(key="min_order_value", value="500", label="Minimalna wartość zlecenia (PLN)"),
        Setting(key="default_overhead_pct", value="0.10", label="Narzut overhead (%)"),
        Setting(key="default_margin_pct", value="0.25", label="Domyślna marża (%)"),
    ]
    for s in defaults:
        # merge = insert jeśli nie istnieje, update jeśli istnieje
        db.merge(s)
    db.commit()
    print(f"✓ {len(defaults)} ustawień systemowych wstawionych")


def seed_price_history_from_xlsx(db: Session):
    """
    Parsuje Lista zleceń usługi.xlsx → price_history.
    Pozwala Claude API w przyszłości sugerować ceny na podstawie historii.
    """
    try:
        import openpyxl
    except ImportError:
        print("⚠ openpyxl nie zainstalowany — pomijam seed price_history")
        return

    if db.query(PriceHistory).count() > 0:
        return

    imported = 0
    for fname in ["Lista zleceń usługi.xlsx", "Lista zleceń usługi_1.xlsx"]:
        fpath = os.path.join(CENY_DIR, fname)
        if not os.path.exists(fpath):
            print(f"⚠ Plik nie znaleziony: {fpath}")
            continue

        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        ws = wb.active

        # Lista zleceń usługi.xlsx ma nagłówki w WIERSZU 3 (nie 1!)
        # Szukamy pierwszego wiersza który zawiera "klient" lub "firma" — elastycznie
        header_row_num = 1
        for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
            vals = [str(v).lower() for v in r if v]
            if any("klient" in v or "firma" in v or "zleceni" in v for v in vals):
                break
            header_row_num += 1

        header_row = next(ws.iter_rows(min_row=header_row_num, max_row=header_row_num, values_only=True))
        headers = {str(v).strip().lower() if v else "": i for i, v in enumerate(header_row)}

        def find_col(*keywords):
            """Znajdź indeks kolumny po częściowym dopasowaniu nazwy (case-insensitive)."""
            for kw in keywords:
                for h, i in headers.items():
                    if kw.lower() in h:
                        return i
            return None

        col_client   = find_col("firma", "klient")
        col_type     = find_col("rodzaj", "wykonania usługi", "typ pracy")
        col_price    = find_col("wycena", "cena", "netto", "wartość")
        col_date     = find_col("data wykonania", "data")
        col_material = find_col("materiał", "material")

        print(f"  Kolumny: klient={col_client}, typ={col_type}, cena={col_price}, data={col_date}")

        for row in ws.iter_rows(min_row=header_row_num + 1, values_only=True):
            if not row or all(v is None for v in row):
                continue
            try:
                client     = str(row[col_client]).strip() if col_client is not None and row[col_client] else ""
                order_type = str(row[col_type]).strip()   if col_type is not None and row[col_type] else "nieznany"
                material   = str(row[col_material]).strip() if col_material is not None and row[col_material] else ""
                order_date = row[col_date] if col_date is not None else None

                # Parsowanie ceny — obsługuje: "544 netto", "1 200,00 zł", "1200.50"
                raw_price = str(row[col_price]) if col_price is not None and row[col_price] else "0"
                clean     = re.sub(r"[^\d.,]", "", raw_price).replace(",", ".")
                price     = float(clean) if clean else 0.0

                if isinstance(order_date, str):
                    from datetime import datetime as dt
                    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d-%m-%Y"):
                        try:
                            order_date = dt.strptime(order_date, fmt).date(); break
                        except ValueError:
                            pass
                    else:
                        order_date = None

                if not client and price == 0.0:
                    continue  # pomiń puste wiersze

                record = PriceHistory(
                    order_type             = order_type,
                    total_price_historical = price,
                    parameters_json        = {"material": material},
                    source                 = fname,
                    order_date             = order_date,
                    client                 = client,
                )
                db.add(record)
                imported += 1
            except (TypeError, ValueError, IndexError):
                continue

        db.commit()
        print(f"✓ {imported} rekordów price_history z {fname}")


def seed_approved_materials(db: Session):
    """Whitelist zatwierdzonych materiałów — 8 podstawowych dla RCM."""
    if db.query(ApprovedMaterial).count() > 0:
        return
    materials = [
        ApprovedMaterial(name="S235",             category="stal",       default_rate_pln_kg=15),
        ApprovedMaterial(name="S355",             category="stal",       default_rate_pln_kg=16),
        ApprovedMaterial(name="S235JR",           category="stal",       default_rate_pln_kg=15),
        ApprovedMaterial(name="stal nierdzewna",  category="nierdzewka", default_rate_pln_kg=22, notes="tylko po uzgodnieniu z Dyrektorem"),
        ApprovedMaterial(name="nierdzewka 304",   category="nierdzewka", default_rate_pln_kg=22),
        ApprovedMaterial(name="żeliwo",           category="żeliwo",     default_rate_pln_kg=18),
        ApprovedMaterial(name="stal konstrukcyjna", category="stal",     default_rate_pln_kg=15),
        ApprovedMaterial(name="P265GH",           category="stal",       default_rate_pln_kg=17, notes="ciśnieniowa"),
    ]
    db.add_all(materials)
    db.commit()
    print(f"✓ {len(materials)} zatwierdzonych materiałów wstawionych")


if __name__ == "__main__":
    engine = init_db()
    with Session(engine) as db:
        seed_users(db)
        seed_constraint_rules(db)
        seed_product_templates(db)
        seed_settings(db)
        seed_approved_materials(db)
        seed_price_history_from_xlsx(db)
    print("\n✅ Seed zakończony. Baza gotowa.")
